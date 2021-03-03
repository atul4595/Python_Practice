from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import pandas as pd
import openpyxl 
import time
import re
wb = openpyxl.Workbook()
sheet = wb.active 
browser = webdriver.Chrome()

try:
    browser.fullscreen_window()
    browser.maximize_window()
except Exception:
    print("Chrome is not able to maximize")

try:
    link = 'https://www.trip.com/'
    browser.get(link)
    time.sleep(5)
except Exception:
    print("URL is wrong")

try:
    browser.find_element_by_xpath("//input[@id='hotels-destination']").clear()
    browser.find_element_by_xpath("//input[@id='hotels-destination']").send_keys('Beijing')
    time.sleep(1.5)
except Exception:
    print("Name of the Place enter wrong")

try:
    d=browser.find_element_by_xpath("//div[@class='time-tab checkin']/input")
    time.sleep(1.5)
    browser.execute_script('arguments[0].removeAttribute("readonly","readonly")', d)
    time.sleep(1.5)
    d.clear()
    time.sleep(1.5)

    a=browser.find_element_by_xpath("//h3[@class='c-calendar-month__title']").text

    while a != "Nov 2020":
        browser.find_element_by_xpath("//span[@class='c-calendar-icon-next']").click()
        a=browser.find_element_by_xpath("//h3[@class='c-calendar-month__title']").text

    date_C= browser.find_element_by_xpath("//h3[@class='c-calendar-month__title']") 
    event = browser.find_elements_by_class_name('c-calendar-month__week')

    for items in event:
        if(a==True):
            break
        li_c = items.find_elements_by_tag_name('li')

        for dat_c in li_c:
            if dat_c.text=="18":
                dat_c.click()
                a=True
                break

    time.sleep(1.5)
except Exception:
    print("Checkin Date is wrong")


try:
    d1=browser.find_element_by_xpath("//div[@class='time-tab checkout']/input")
    time.sleep(1.5)
    browser.execute_script('arguments[0].removeAttribute("readonly","readonly")', d1)
    time.sleep(1.5)
    d1.clear()
    time.sleep(1.5)
    a1=browser.find_element_by_xpath("//h3[@class='c-calendar-month__title']").text

    while a1 != "Nov 2020":
        browser.find_element_by_xpath("//span[@class='c-calendar-icon-next']").click()
        a1=browser.find_element_by_xpath("//h3[@class='c-calendar-month__title']").text

    date_C= browser.find_element_by_xpath("//h3[@class='c-calendar-month__title']") 
    event = browser.find_elements_by_class_name('c-calendar-month__week')

    for items in event:
        if(a1==True):
            break

        li_c = items.find_elements_by_tag_name('li')
        
        for dat_c in li_c:
            if dat_c.text=="25":
                dat_c.click()
                time.sleep(1.5)
                a1=True
                break
except Exception:
    print("Checkout date is wrong")
            
browser.find_element_by_xpath("//div[@class='child-kid'][2]/div[1]/span[3]/i[1]").click()
time.sleep(1.5)
browser.find_element_by_xpath("//div[@class='search-btn-wrap'][1]/i[1]").click()
time.sleep(1.5)

Search_More=""

while Search_More!="Search More Hotels":
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1.5)

    try:
        Search_More=browser.find_element_by_xpath("//div[@class='list-btn-more']/p[1]").text
    except Exception:
        Search_More=""

Total_Hotels=browser.find_element_by_xpath("//div[@class='filter-title clearfix']/h3[1]").text
Total_Hotels=Total_Hotels.replace(",","")
numbers = int(re.findall('[0-9]+', Total_Hotels)[0])
Show_l=len(browser.find_elements_by_class_name('list-card-title'))
# For Current scenario i am extracting data for 30 Hotel if you want to extract more hotel just comment below line
numbers=30

while Show_l != numbers:
    browser.find_element_by_xpath("//div[@class='list-btn-more']/p[1]").click()
    time.sleep(1.5)
    Show_l=len(browser.find_elements_by_class_name('list-card-title'))

# browser.find_element_by_tag_name('body').send_keys(Keys.CONTROL + Keys.HOME)
browser.execute_script("window.scrollTo(0, 0);")
time.sleep(1.5)
event_C = browser.find_elements_by_class_name('list-card-title')

try:
    sheet.cell(row= 1 , column = 1).value="Hotel_Name"
    sheet.cell(row= 1 , column = 2).value="Room_Name"
    sheet.cell(row= 1 , column = 3).value="Bed_type"
    sheet.cell(row= 1 , column = 4).value="Rating"
    sheet.cell(row= 1 , column = 5).value="Price"
except Exception:
    print("Can't able to give Header")

row_inc=2

for items_C in event_C:
    # browser.switch_to.window(browser.window_handles[0])
    items_C.click()
    browser.switch_to.window(browser.window_handles[1])
    Hotel_Name=browser.find_element_by_xpath("//section[contains(@class, 'detail-baseinfo_title')]/h1[1]").text
    time.sleep(1.5)
    Rating=len(browser.find_elements_by_xpath("//section[contains(@class, 'detail-baseinfo_title')]/i"))
    time.sleep(1.5)
    Rooms = browser.find_elements_by_class_name('roomlist-baseroom-card')
    time.sleep(1.5)

    for Room in Rooms:
        Room_Name=Room.find_element_by_class_name('roomname').text
        Room_D=Room.find_elements_by_class_name('salecard-frame')
        for Room_1_D in Room_D:
            bed=Room_1_D.find_element_by_class_name('salecard-bedfacility').text
            price=Room_1_D.find_element_by_class_name('note').text
            sheet.cell(row= row_inc , column = 1).value=Hotel_Name
            sheet.cell(row= row_inc , column = 2).value=Room_Name
            sheet.cell(row= row_inc , column = 3).value=bed
            sheet.cell(row= row_inc , column = 4).value=Rating
            sheet.cell(row= row_inc , column = 5).value=price
            row_inc = row_inc + 1

    browser.close()
    browser.switch_to.window(browser.window_handles[0])
wb.save("Output.xlsx")
browser.close()